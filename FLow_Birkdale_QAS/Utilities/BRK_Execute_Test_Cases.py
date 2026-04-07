"""
================================================================================
  Synovia Flow -- TSS Birkdale Test Case Executor
  Licensed Component: Synovia Digital Ltd
================================================================================

  Product:      Synovia Flow (Customs Declaration Management)
  Module:       TSS Birkdale Test Case Executor
  Version:      1.0.0
  Database:     Fusion_TSS
  Schema:       BKD (Birkdale)
  API:          TSS Declaration API v2.9.4 (TEST)

  Description:
  ------------
  Reads test case payloads from BRK_Test_Cases.xlsx, executes each
  against the TSS TEST API in the correct sequence, and writes a
  comprehensive results workbook with every call, response, and issue.

  Input:   D:\\TSS_Madrid\\BRK_Test_Cases.xlsx
  Output:  D:\\TSS_Madrid\\Birkdale\\brk_test_results_<timestamp>.xlsx
  Log:     BKD.ApiLog in Fusion_TSS

  Execution sequence per ENS test case:
    Step 1: POST /headers              (create ENS header)
    Step 2: POST /consignments         (create consignment, link to ENS)
    Step 3: POST /goods                (create goods, link to consignment)
    Step 4: POST /consignments         (submit consignment)
    Step 5: GET  /consignments         (read back status + MRN)
    Step 6: GET  /simplified_frontier_declarations?consignment_number=
    Step 7: GET  /supplementary_declarations?sfd_number=

  Execution sequence per FFD test case:
    Step 1: POST /full_frontier_declarations  (create FFD)
    Step 2: POST /goods                       (create goods, link to FFD)
    Step 3: GET  /full_frontier_declarations  (read back status)

  Copyright (c) 2026 Synovia Digital Ltd. All rights reserved.
================================================================================
"""

__version__ = '1.0.0'
__product__ = 'Synovia Flow'
__module__  = 'TSS Birkdale Test Case Executor'

import base64, configparser, json, os, sys, time
from datetime import datetime, timezone, timedelta
import pyodbc, requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich import box

con = Console(highlight=False, width=140)

CLIENT_CODE = 'BKD'
CLIENT_NAME = 'Birkdale'
ENV_CODE    = 'TST'
DB_NAME     = 'Fusion_TSS'
INI_PATH    = r'D:\confguration\fusion_TSS.ini'
S           = CLIENT_CODE

RATE_LIMIT  = 0.30
API_TIMEOUT = 30
LOG_BATCH   = 10

TIMESTAMP   = datetime.now().strftime('%Y%m%d_%H%M%S')
INPUT_FILE  = r'D:\TSS_Madrid\BRK_Test_Cases.xlsx'
OUTPUT_DIR  = r'D:\TSS_Madrid\Birkdale'
RESULTS_FILE = os.path.join(OUTPUT_DIR, f'brk_test_results_{TIMESTAMP}.xlsx')

# Styling
NAVY = '1B2A4A'; BLUE = '2E86AB'; GREEN = '28A745'
AMBER = 'E8A317'; RED = 'C62828'; WHITE = 'FFFFFF'
LIGHT = 'F0F4F8'
hdr_font = Font(name='Arial', bold=True, color=WHITE, size=10)
data_font = Font(name='Arial', size=9)
mono_font = Font(name='Consolas', size=8)
ok_font = Font(name='Arial', size=9, color='28A745')
fail_font = Font(name='Arial', size=9, color='C62828')
alt_fill = PatternFill('solid', fgColor=LIGHT)
bdr = Border(bottom=Side(style='thin', color='CCCCCC'))


def make_conn():
    cfg = configparser.ConfigParser(); cfg.read(INI_PATH)
    d = cfg['database']
    return pyodbc.connect(
        f"DRIVER={d['driver']};SERVER={d['server']};DATABASE={DB_NAME};"
        f"UID={d['user']};PWD={d['password']};"
        f"Encrypt={d.get('encrypt','yes')};"
        f"TrustServerCertificate={d.get('trust_server_certificate','no')};",
        autocommit=False)

def query(sql, params=None):
    conn = make_conn(); cur = conn.cursor()
    cur.execute(sql, params or [])
    cols = [c[0] for c in cur.description] if cur.description else []
    rows = cur.fetchall(); conn.close()
    return [dict(zip(cols, r)) for r in rows]

def load_credentials():
    rows = query("""
        SELECT e.base_url, cr.tss_username, cr.tss_password
        FROM CFG.Credentials cr JOIN CFG.Environments e ON e.env_code=cr.env_code
        WHERE cr.client_code=? AND cr.env_code=? AND cr.active=1
    """, [CLIENT_CODE, ENV_CODE])
    if not rows:
        con.print(f'[red]No active {ENV_CODE} creds for {CLIENT_CODE}[/red]'); sys.exit(1)
    return rows[0]


class ApiLogger:
    def __init__(self):
        self.buffer = []; self.total_flushed = 0
    def log(self, dt, ref, ep, params, http, raw, ms, method='GET', notes=''):
        self.buffer.append((
            (dt or '')[:50], 'WRITE' if method=='POST' else 'READ',
            (ref or '')[:200], None, CLIENT_CODE, method,
            (ep or '')[:500], json.dumps(params, separators=(',',':'))[:4000],
            http, 'OK' if http in (200,201) else 'FAIL',
            (raw or '')[:500], (raw or '')[:4000], ms,
            '' if http in (200,201) else (raw or '')[:4000],
            (notes or f'Executor v{__version__}')[:200]))
        if len(self.buffer) >= LOG_BATCH: self.flush()
    def flush(self):
        if not self.buffer: return
        try:
            conn = make_conn(); cur = conn.cursor()
            cur.executemany(f"""
                INSERT INTO {S}.ApiLog (declaration_type,call_type,reference,
                    act_as,act_as_customer,http_method,url,request_params,
                    http_status,response_status,response_message,response_json,
                    duration_ms,error_detail,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", self.buffer)
            conn.commit(); conn.close()
            self.total_flushed += len(self.buffer)
        except Exception as e:
            con.print(f'    [dim red]ApiLog: {e}[/dim red]')
        self.buffer.clear()


class TssApi:
    def __init__(self, base_url, username, password, logger):
        self.base_url = base_url.rstrip('/') + '/x_fhmrc_tss_api/v1/tss_api'
        self.logger = logger; self.session = requests.Session()
        b64 = base64.b64encode(f'{username}:{password}'.encode()).decode()
        self.session.headers.update({
            'Accept':'application/json','Content-Type':'application/json',
            'Authorization':f'Basic {b64}'})
        self.total_calls = 0
        self.call_log = []  # [(step, tc, method, endpoint, params, http, ref_returned, response_snippet, ms, error)]

    def _call(self, method, ep, params_or_payload, step='', tc='', dt='', ref='', notes=''):
        url = f'{self.base_url}/{ep}'
        t0 = time.time()
        snippet = ''; ref_returned = ''; error = ''
        try:
            if method == 'POST':
                r = self.session.post(url, json=params_or_payload, timeout=API_TIMEOUT)
                param_str = json.dumps(params_or_payload, separators=(',',':'))[:120]
            else:
                r = self.session.get(url, params=params_or_payload, timeout=API_TIMEOUT)
                param_str = '&'.join(f'{k}={v}' for k,v in params_or_payload.items() if k!='fields')[:120]

            self.total_calls += 1
            ms = int((time.time()-t0)*1000)
            time.sleep(RATE_LIMIT)

            self.logger.log(dt, ref, ep, params_or_payload,
                           r.status_code, r.text[:4000], ms,
                           method=method, notes=notes)

            snippet = r.text[:500]
            if r.status_code == 200:
                result = r.json().get('result')
                if isinstance(result, dict):
                    ref_returned = result.get('reference', '') or result.get('sfd_number', '') or result.get('sup_dec_number', '')
                    status = result.get('status', '') or result.get('process_message', '')
                elif isinstance(result, list):
                    ref_returned = str(len(result)) + ' items'
                con.print(f'    [{("green" if r.status_code==200 else "red")}]{method} /{ep}  '
                          f'HTTP {r.status_code}  ref={ref_returned}  {ms}ms[/]')
            else:
                try:
                    body = r.json()
                    error = (body.get('result',{}).get('process_message','')
                             or body.get('error',{}).get('message','')
                             or r.text[:200])
                except: error = r.text[:200]
                con.print(f'    [red]{method} /{ep}  HTTP {r.status_code}  {error[:80]}  {ms}ms[/red]')

            self.call_log.append((step, tc, method, ep, param_str, r.status_code,
                                  ref_returned, snippet[:300], ms, error))
            return r.status_code, r.json().get('result') if r.status_code==200 else None, r.text, ms

        except Exception as e:
            ms = int((time.time()-t0)*1000)
            self.total_calls += 1
            error = str(e)[:200]
            self.logger.log(dt, ref, ep, params_or_payload, 0, error, ms, method=method, notes=notes)
            self.call_log.append((step, tc, method, ep, str(params_or_payload)[:120], 0, '', '', ms, error))
            con.print(f'    [red]{method} /{ep}  EXCEPTION  {error[:80]}[/red]')
            return 0, None, error, ms


def sget(d, k, default=''):
    return d.get(k, default) if isinstance(d, dict) else default


def write_results_sheet(ws, columns, records, tab_color, header_fill):
    ws.sheet_properties.tabColor = tab_color
    for ci, (name, width) in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = hdr_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    for ri, rec in enumerate(records, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(rec, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = data_font; c.border = bdr
            if fill: c.fill = fill
            # Colour HTTP status
            if ci == 6 and isinstance(val, int):
                c.font = ok_font if val == 200 else fail_font
    ws.auto_filter.ref = f'A1:{get_column_letter(len(columns))}1'
    ws.freeze_panes = 'A2'


def main():
    t0 = time.time()

    con.print(Panel.fit(
        f'[bold yellow]{__product__}[/bold yellow]  |  '
        f'[bold white]{__module__}[/bold white]  v{__version__}\n'
        f'[bold cyan]{CLIENT_NAME}[/bold cyan]  |  '
        f'[dim]{CLIENT_CODE}  |  {ENV_CODE}  |  {DB_NAME}  |  '
        f'{datetime.now(timezone.utc):%Y-%m-%d %H:%M:%S} UTC[/dim]\n'
        f'[dim]Input: {INPUT_FILE}[/dim]',
        border_style='blue', padding=(0, 2)))

    # ── Preflight ─────────────────────────────────────────────
    con.print(); con.rule('[bold cyan]Preflight[/bold cyan]'); con.print()

    if not os.path.exists(INPUT_FILE):
        con.print(f'  [red]Input not found: {INPUT_FILE}[/red]'); return
    con.print(f'  Input:  [green]OK[/green]  {INPUT_FILE}')

    if not os.path.exists(INI_PATH):
        con.print(f'  [red]INI missing[/red]'); return

    try:
        lc = query(f"SELECT COUNT(*) AS c FROM {S}.ApiLog")[0]['c']
        con.print(f'  ApiLog: [green]OK[/green]  ({lc} rows)')
    except:
        con.print(f'  ApiLog: [red]MISSING[/red]'); return

    creds = load_credentials()
    con.print(f'  API:    [green]OK[/green]  {creds["tss_username"]}')
    con.print(f'  Base:   [dim]{creds["base_url"]}[/dim]')
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    logger = ApiLogger()
    api = TssApi(creds['base_url'], creds['tss_username'], creds['tss_password'], logger)

    # ── Load Excel ────────────────────────────────────────────
    con.print(); con.rule('[bold cyan]Load Test Cases[/bold cyan]'); con.print()

    wb_in = load_workbook(INPUT_FILE, data_only=True)
    con.print(f'  Sheets: {wb_in.sheetnames}')

    # Read ENS headers
    ens_cases = []
    if 'TC1_ENS_Header' in wb_in.sheetnames:
        ws = wb_in['TC1_ENS_Header']
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                d = dict(zip(headers, row))
                ens_cases.append(d)
        con.print(f'  ENS Headers: [cyan]{len(ens_cases)}[/cyan] test cases')

    # Read Consignments
    cons_cases = []
    if 'TC1_Consignment' in wb_in.sheetnames:
        ws = wb_in['TC1_Consignment']
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                cons_cases.append(dict(zip(headers, row)))
        con.print(f'  Consignments: [cyan]{len(cons_cases)}[/cyan]')

    # Read Goods
    goods_cases = []
    if 'TC1_Goods' in wb_in.sheetnames:
        ws = wb_in['TC1_Goods']
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                goods_cases.append(dict(zip(headers, row)))
        con.print(f'  Goods Items: [cyan]{len(goods_cases)}[/cyan]')

    # Read FFDs
    ffd_cases = []
    if 'TC2_FFD' in wb_in.sheetnames:
        ws = wb_in['TC2_FFD']
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ffd_cases.append(dict(zip(headers, row)))
        con.print(f'  FFDs: [cyan]{len(ffd_cases)}[/cyan]')

    # Read FFD Goods
    ffd_goods_cases = []
    if 'TC2_FFD_Goods' in wb_in.sheetnames:
        ws = wb_in['TC2_FFD_Goods']
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ffd_goods_cases.append(dict(zip(headers, row)))
        con.print(f'  FFD Goods: [cyan]{len(ffd_goods_cases)}[/cyan]')

    # ── Compute arrival time ──────────────────────────────────
    # Near future for submission (5 min from now GMT)
    arrival_future = (datetime.now(timezone.utc) + timedelta(minutes=5)).strftime('%d/%m/%Y %H:%M:%S')
    # Past for arrived testing
    arrival_past = (datetime.now(timezone.utc) - timedelta(hours=1)).strftime('%d/%m/%Y %H:%M:%S')
    con.print(f'  Arrival (future): {arrival_future}')
    con.print(f'  Arrival (past):   {arrival_past}')

    # ══════════════════════════════════════════════════════════
    #  Execute TC1: ENS -> Consignment -> Goods -> Submit
    # ══════════════════════════════════════════════════════════
    con.print(); con.rule('[bold green]Execute TC1: ENS Simplified Procedure[/bold green]'); con.print()

    tc1_results = {}  # {test_case: {ens_ref, cons_ref, goods_ids, sfd_ref, sd_ref, ...}}

    for ens in ens_cases:
        tc = ens['test_case']
        con.print(f'\n  [bold yellow]━━ {tc} ━━[/bold yellow]')
        tc1_results[tc] = {'ens_ref':'','cons_ref':'','goods_ids':[],'sfd_ref':'','sd_ref':'','errors':[]}

        # Parse the JSON payload from Excel
        try:
            payload = json.loads(ens.get('json_payload','{}'))
        except:
            payload = {}

        # Set the arrival time
        payload['arrival_date_time'] = arrival_future

        # Step 1: Create ENS Header
        http, result, raw, ms = api._call('POST', 'headers', payload,
            step='1-Create-ENS', tc=tc, dt='ENS_CREATE', notes=f'{tc} header')
        if http == 200 and result:
            ens_ref = sget(result, 'reference')
            tc1_results[tc]['ens_ref'] = ens_ref
            con.print(f'  [green]ENS: {ens_ref}[/green]')

            # Find matching consignment
            matching_cons = [c for c in cons_cases if c['test_case'] == tc]
            for mc in matching_cons:
                try:
                    cons_payload = json.loads(mc.get('json_payload','{}'))
                except:
                    cons_payload = {}
                cons_payload['declaration_number'] = ens_ref

                # Step 2: Create Consignment
                http2, result2, raw2, ms2 = api._call('POST', 'consignments', cons_payload,
                    step='2-Create-Cons', tc=tc, dt='CONS_CREATE', ref=ens_ref, notes=f'{tc} consignment')
                if http2 == 200 and result2:
                    cons_ref = sget(result2, 'reference')
                    tc1_results[tc]['cons_ref'] = cons_ref
                    con.print(f'  [green]Consignment: {cons_ref}[/green]')

                    # Find matching goods
                    tc_prefix = tc  # exact match or startswith
                    matching_goods = [g for g in goods_cases
                                      if g['test_case'].startswith(tc.replace('_Controlled','').replace('_Uncontrolled','').replace('_MultiGoods',''))
                                      or g['test_case'].startswith(tc)]
                    # Better: match by test_case prefix
                    matching_goods = [g for g in goods_cases if g['test_case'].startswith(tc)]

                    for gi, mg in enumerate(matching_goods, 1):
                        try:
                            goods_payload = json.loads(mg.get('json_payload','{}'))
                        except:
                            goods_payload = {}
                        goods_payload['consignment_number'] = cons_ref

                        # Step 3: Create Goods
                        http3, result3, raw3, ms3 = api._call('POST', 'goods', goods_payload,
                            step=f'3-Create-Goods-L{gi}', tc=tc, dt='GOODS_CREATE',
                            ref=cons_ref, notes=f'{tc} goods {gi}')
                        if http3 == 200 and result3:
                            gid = sget(result3, 'reference')
                            tc1_results[tc]['goods_ids'].append(gid)
                            con.print(f'  [green]Goods L{gi}: {gid}[/green]')
                        else:
                            tc1_results[tc]['errors'].append(f'Goods L{gi}: HTTP {http3}')

                    # Step 4: Submit Consignment
                    submit_payload = {'op_type': 'submit', 'consignment_number': cons_ref}
                    http4, result4, raw4, ms4 = api._call('POST', 'consignments', submit_payload,
                        step='4-Submit-Cons', tc=tc, dt='CONS_SUBMIT', ref=cons_ref, notes=f'{tc} submit')
                    if http4 == 200 and result4:
                        sub_status = sget(result4, 'status', sget(result4, 'process_message', ''))
                        con.print(f'  [green]Submitted: {sub_status}[/green]')
                    else:
                        tc1_results[tc]['errors'].append(f'Submit: HTTP {http4}')

                    # Step 5: Read back consignment
                    time.sleep(1)
                    http5, result5, raw5, ms5 = api._call('GET', 'consignments',
                        {'reference': cons_ref, 'fields': 'status,declaration_number,movement_reference_number,controlled_goods'},
                        step='5-Read-Cons', tc=tc, dt='CONS_READ', ref=cons_ref, notes=f'{tc} verify')
                    if http5 == 200 and result5:
                        con.print(f'  Status: {sget(result5,"status")}  '
                                  f'MRN: {sget(result5,"movement_reference_number")}')

                    # Step 6: SFD Lookup
                    http6, result6, raw6, ms6 = api._call('GET', 'simplified_frontier_declarations',
                        {'consignment_number': cons_ref},
                        step='6-SFD-Lookup', tc=tc, dt='SFD_LOOKUP', ref=cons_ref, notes=f'{tc} SFD lookup')
                    if http6 == 200 and result6:
                        sfd_num = sget(result6, 'sfd_number')
                        if sfd_num:
                            tc1_results[tc]['sfd_ref'] = sfd_num
                            con.print(f'  [green]SFD: {sfd_num}[/green]')

                            # Step 7: SD Lookup
                            http7, result7, raw7, ms7 = api._call('GET', 'supplementary_declarations',
                                {'sfd_number': sfd_num},
                                step='7-SD-Lookup', tc=tc, dt='SD_LOOKUP', ref=sfd_num, notes=f'{tc} SD lookup')
                            if http7 == 200 and result7:
                                sd_num = sget(result7, 'sup_dec_number')
                                if sd_num:
                                    tc1_results[tc]['sd_ref'] = sd_num
                                    con.print(f'  [green]SD: {sd_num}[/green]')
                                else:
                                    con.print(f'  [yellow]No SD yet (may need arrival)[/yellow]')
                        else:
                            con.print(f'  [yellow]No SFD yet[/yellow]')
                else:
                    tc1_results[tc]['errors'].append(f'Consignment: HTTP {http2}')
        else:
            tc1_results[tc]['errors'].append(f'Header: HTTP {http}')

        logger.flush()

    # ══════════════════════════════════════════════════════════
    #  Execute TC2: FFD
    # ══════════════════════════════════════════════════════════
    con.print(); con.rule('[bold green]Execute TC2: Full Frontier Declarations[/bold green]'); con.print()

    tc2_results = {}

    for ffd in ffd_cases:
        tc = ffd['test_case']
        con.print(f'\n  [bold yellow]━━ {tc} ━━[/bold yellow]')
        tc2_results[tc] = {'ffd_ref':'','goods_ids':[],'errors':[]}

        try:
            payload = json.loads(ffd.get('json_payload','{}'))
        except:
            payload = {}
        payload['arrival_date_time'] = arrival_future

        # Step 1: Create FFD
        http, result, raw, ms = api._call('POST', 'full_frontier_declarations', payload,
            step='1-Create-FFD', tc=tc, dt='FFD_CREATE', notes=f'{tc} FFD')
        if http == 200 and result:
            ffd_ref = sget(result, 'reference')
            tc2_results[tc]['ffd_ref'] = ffd_ref
            con.print(f'  [green]FFD: {ffd_ref}[/green]')

            # FFD Goods
            matching_goods = [g for g in ffd_goods_cases if g['test_case'] == tc]
            for gi, mg in enumerate(matching_goods, 1):
                try:
                    goods_payload = json.loads(mg.get('json_payload','{}'))
                except:
                    goods_payload = {}
                goods_payload['consignment_number'] = ffd_ref

                http2, result2, raw2, ms2 = api._call('POST', 'goods', goods_payload,
                    step=f'2-Create-FFD-Goods-L{gi}', tc=tc, dt='FFD_GOODS_CREATE',
                    ref=ffd_ref, notes=f'{tc} goods {gi}')
                if http2 == 200 and result2:
                    gid = sget(result2, 'reference')
                    tc2_results[tc]['goods_ids'].append(gid)
                    con.print(f'  [green]FFD Goods L{gi}: {gid}[/green]')
                else:
                    tc2_results[tc]['errors'].append(f'FFD Goods: HTTP {http2}')

            # Read back FFD
            http3, result3, raw3, ms3 = api._call('GET', 'full_frontier_declarations',
                {'reference': ffd_ref, 'fields': 'status,importer_eori,movement_type'},
                step='3-Read-FFD', tc=tc, dt='FFD_READ', ref=ffd_ref, notes=f'{tc} verify')
        else:
            tc2_results[tc]['errors'].append(f'FFD: HTTP {http}')

        logger.flush()

    # ══════════════════════════════════════════════════════════
    #  Write Results Excel
    # ══════════════════════════════════════════════════════════
    elapsed = time.time() - t0
    logger.flush()

    con.print(); con.rule('[bold cyan]Write Results Excel[/bold cyan]'); con.print()

    wb_out = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws_sum = wb_out.active; ws_sum.title = 'Summary'
    ws_sum.sheet_properties.tabColor = NAVY
    ws_sum.column_dimensions['A'].width = 30; ws_sum.column_dimensions['B'].width = 50

    sum_rows = [
        (f'{__product__} -- {__module__} v{__version__}', ''),
        ('', ''),
        ('Generated', datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')),
        ('Input File', INPUT_FILE),
        ('Client', f'{CLIENT_NAME} ({CLIENT_CODE})'),
        ('Environment', ENV_CODE),
        ('API Base', creds['base_url']),
        ('Total API Calls', str(api.total_calls)),
        ('Runtime', f'{elapsed:.1f}s'),
        ('', ''),
    ]
    # TC1 results
    for tc, r in tc1_results.items():
        sum_rows.append((f'TC1: {tc}', ''))
        sum_rows.append(('  ENS Header', r['ens_ref'] or 'FAILED'))
        sum_rows.append(('  Consignment', r['cons_ref'] or 'FAILED'))
        sum_rows.append(('  Goods Items', str(len(r['goods_ids']))))
        sum_rows.append(('  SFD Generated', r['sfd_ref'] or 'Not yet'))
        sum_rows.append(('  SD Generated', r['sd_ref'] or 'Not yet'))
        sum_rows.append(('  Errors', str(len(r['errors'])) if r['errors'] else 'None'))
        sum_rows.append(('', ''))

    for tc, r in tc2_results.items():
        sum_rows.append((f'TC2: {tc}', ''))
        sum_rows.append(('  FFD', r['ffd_ref'] or 'FAILED'))
        sum_rows.append(('  Goods Items', str(len(r['goods_ids']))))
        sum_rows.append(('  Errors', str(len(r['errors'])) if r['errors'] else 'None'))
        sum_rows.append(('', ''))

    for ri, (k, v) in enumerate(sum_rows, 1):
        c1 = ws_sum.cell(row=ri, column=1, value=k)
        c2 = ws_sum.cell(row=ri, column=2, value=v)
        c1.font = Font(name='Arial', bold=True, size=14 if ri==1 else 9, color=NAVY)
        c2.font = data_font
        if 'FAILED' in str(v): c2.font = fail_font

    # ── Sheet 2: All API Calls ────────────────────────────────
    ws_calls = wb_out.create_sheet('API_Calls')
    call_cols = [
        ('call_num',8), ('step',22), ('test_case',25), ('method',8),
        ('endpoint',35), ('http_status',10), ('ref_returned',25),
        ('duration_ms',10), ('error',50), ('request_params',60),
        ('response_snippet',80),
    ]
    call_recs = []
    for ci, (step, tc, method, ep, params, http, ref_ret, snippet, ms, err) in enumerate(api.call_log, 1):
        call_recs.append([ci, step, tc, method, ep, http, ref_ret, ms, err, params, snippet])
    write_results_sheet(ws_calls, call_cols, call_recs, BLUE, PatternFill('solid', fgColor=BLUE))

    # ── Sheet 3: Errors Only ──────────────────────────────────
    ws_errs = wb_out.create_sheet('Errors')
    err_recs = [r for r in call_recs if r[5] != 200]
    write_results_sheet(ws_errs, call_cols, err_recs, RED, PatternFill('solid', fgColor=RED))

    # ── Sheet 4: Created References ───────────────────────────
    ws_refs = wb_out.create_sheet('Created_Refs')
    ref_cols = [
        ('test_case',25), ('type',18), ('reference',30), ('status',20),
    ]
    ref_recs = []
    for tc, r in tc1_results.items():
        if r['ens_ref']: ref_recs.append([tc, 'ENS Header', r['ens_ref'], 'Created'])
        if r['cons_ref']: ref_recs.append([tc, 'Consignment', r['cons_ref'], 'Submitted'])
        for gid in r['goods_ids']: ref_recs.append([tc, 'Goods', gid, 'Created'])
        if r['sfd_ref']: ref_recs.append([tc, 'SFD (auto)', r['sfd_ref'], 'Auto-generated'])
        if r['sd_ref']: ref_recs.append([tc, 'SD (auto)', r['sd_ref'], 'Auto-generated'])
    for tc, r in tc2_results.items():
        if r['ffd_ref']: ref_recs.append([tc, 'FFD', r['ffd_ref'], 'Created'])
        for gid in r['goods_ids']: ref_recs.append([tc, 'FFD Goods', gid, 'Created'])
    write_results_sheet(ws_refs, ref_cols, ref_recs, GREEN, PatternFill('solid', fgColor=GREEN))

    # ── Sheet 5: Test Case Issues ─────────────────────────────
    ws_issues = wb_out.create_sheet('Issues')
    issue_cols = [('test_case',25), ('issue',80)]
    issue_recs = []
    for tc, r in tc1_results.items():
        for e in r['errors']: issue_recs.append([tc, e])
    for tc, r in tc2_results.items():
        for e in r['errors']: issue_recs.append([tc, e])
    if not issue_recs: issue_recs.append(['', 'No issues — all test cases passed'])
    write_results_sheet(ws_issues, issue_cols, issue_recs, AMBER, PatternFill('solid', fgColor=AMBER))

    wb_out.save(RESULTS_FILE)
    con.print(f'  Results: [green]{RESULTS_FILE}[/green]')

    # ── Console Summary ───────────────────────────────────────
    con.print(); con.rule('[bold yellow]Execution Complete[/bold yellow]'); con.print()

    tbl = Table(box=box.ROUNDED,
        title=f'[bold]{CLIENT_NAME} -- Test Case Results[/bold]',
        border_style='green')
    tbl.add_column('Test Case', style='cyan', min_width=25)
    tbl.add_column('ENS', justify='center')
    tbl.add_column('Cons', justify='center')
    tbl.add_column('Goods', justify='center')
    tbl.add_column('SFD', justify='center')
    tbl.add_column('SD', justify='center')
    tbl.add_column('FFD', justify='center')
    tbl.add_column('Errors', justify='center', style='red')

    for tc, r in tc1_results.items():
        tbl.add_row(tc,
            '[green]OK[/green]' if r['ens_ref'] else '[red]FAIL[/red]',
            '[green]OK[/green]' if r['cons_ref'] else '[red]FAIL[/red]',
            str(len(r['goods_ids'])) if r['goods_ids'] else '[red]0[/red]',
            '[green]'+r['sfd_ref'][:12]+'[/green]' if r['sfd_ref'] else '[yellow]pending[/yellow]',
            '[green]'+r['sd_ref'][:12]+'[/green]' if r['sd_ref'] else '[yellow]pending[/yellow]',
            '-',
            str(len(r['errors'])) if r['errors'] else '[green]0[/green]')
    for tc, r in tc2_results.items():
        tbl.add_row(tc, '-', '-',
            str(len(r['goods_ids'])) if r['goods_ids'] else '[red]0[/red]',
            '-', '-',
            '[green]'+r['ffd_ref'][:12]+'[/green]' if r['ffd_ref'] else '[red]FAIL[/red]',
            str(len(r['errors'])) if r['errors'] else '[green]0[/green]')

    tbl.add_row('')
    tbl.add_row('[bold]Total API Calls[/bold]', '', '', f'[bold]{api.total_calls}[/bold]', '', '', '', '')
    tbl.add_row('Runtime', '', '', f'{elapsed:.0f}s', '', '', '', '')
    tbl.add_row('Results', '', '', RESULTS_FILE, '', '', '', '')
    con.print(tbl)

    con.print()
    con.print(f'  [dim]{__product__} v{__version__} -- {__module__} -- '
              f'{CLIENT_NAME} -- Synovia Digital Ltd[/dim]')
    con.print()


if __name__ == '__main__':
    main()
